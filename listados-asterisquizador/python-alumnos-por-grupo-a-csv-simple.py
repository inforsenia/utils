#!/usr/bin/python3

# -*- coding: utf-8 -*-

# This script is licensed under 
# GPL v3 or Higher.

# Author: Angel Berlanas Vicente
# eMail : a.berlanasvicente@edu.gva.es

 # -*- coding: utf-8 -*-


import sys
import os
import shutil
import glob

import codecs


from openpyxl import Workbook
from openpyxl import load_workbook

# Posiciones de la hoja Excel
pos_nombre = 1
pos_apellido = 2
pos_apellido_dos = 3
pos_grupo = 12

# Rutas a las carpetas destino de los ficheros

ruta_des = "./generados-csv-simple/"

ruta_des = "./generados-csv/"


# Funciones Chachis
def lopdize(aux):
	apelFinal = ""

	for auxp in aux.split(" "):
		auxf = auxp[0:2]

		sasteriscos = ""
		nasteriscos = len(auxp)-2
		for i in range(nasteriscos):
			sasteriscos += "*"

		#apelFinal+=auxf.capitalize()+sasteriscos+" "
		apelFinal+=auxf+sasteriscos+" "

	return(apelFinal)



listFile = sys.argv[1]
wb = load_workbook(listFile)

# Limpiamos los ficheros de ejecuciones anteriores
files = glob.glob(ruta_des+'/*.csv', recursive=True)

for f in files:
	try:
		os.remove(f)
	except OSError as e:
		print("Error: %s : %s" % (f, e.strerror))

os.makedirs(ruta_des, exist_ok=True)

for sheet_name in wb.sheetnames:

	sheet = wb[sheet_name]
	for row in sheet.iter_rows():
		try:
			grupo = row[pos_grupo].value
			
		except Exception as e:
			print(str(e))
			pass

		if grupo != None:
			#print( " * Alumno/a del Grupo : " + grupo)
			#print( "    - Nombre : " + row[pos_nombre].value)
			#print( "    - Apellido : " + row[pos_apellido].value +" " + str(row[pos_apellido_dos].value))
			
			# Tratamos posibles nones en los apellidos
			nombre = row[pos_nombre].value
			#apellidos = lopdize(row[pos_apellido].value)
<<<<<<< HEAD
			
=======
>>>>>>> d5c8d4e (Updated)
			# Aqui no aplicamos lopdize
			apellidos = row[pos_apellido].value
			#if row[pos_apellido_dos].value != None:
			#	# apellidos += " "+lopdize(row[pos_apellido_dos].value)
			#	apellidos += ","+row[pos_apellido_dos].value
			
<<<<<<< HEAD
			print("    - Simple : "+apellidos)
=======
			#print("    - Lopdize : "+apellidos)
			print("    - Normal : "+apellidos)
			
			# Escribiendo en el fichero de markdown del grupo
>>>>>>> d5c8d4e (Updated)
			
			# Comprobamos que el fichero del grupo existe
			
			file_grupo = ruta_des+grupo+".csv"
			
			if not os.path.exists(file_grupo):
				print(" * [ GRUPO ] : " + file_grupo + " --> Creandolo...")
<<<<<<< HEAD
				f = open(file_grupo, "a", encoding='utf-8')
				f.write("Nombre,Apellidos 2\n")
=======
				f = open(file_grupo, "a")
				f.write("Nombre,Apellidos\n")
>>>>>>> d5c8d4e (Updated)
				f.close()
				
			if os.path.exists(file_grupo):
				
				# Abrimos el fichero
<<<<<<< HEAD
				f = open(file_grupo, "a", encoding='utf-8')
				f.write(nombre+","+apellidos+"\n")
				f.close()


			file_grupo_simple = ruta_des+grupo+"_simple.csv"

			if not os.path.exists(file_grupo_simple):
				print(" * [ GRUPO ] : " + file_grupo_simple + " --> Creandolo...")
				f = open(file_grupo_simple, "a", encoding='utf-8')
				f.write("Nombre y Apellidos 2\n")
				f.close()
				
			if os.path.exists(file_grupo_simple):
				
				# Abrimos el fichero
				f = open(file_grupo_simple, "a", encoding='utf-8')
				f.write(nombre+" "+apellidos+"\n")
				f.close()





=======
				f = open(file_grupo, "a")
				f.write(nombre+","+apellidos+"\n")
				f.close()
>>>>>>> d5c8d4e (Updated)
				
		else:
			print(" * [ Error ] : "+ str(grupo) +" no encontrado")
